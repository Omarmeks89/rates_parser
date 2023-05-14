import typing
import abc
import functools
import collections

import openpyxl as oppxl

from .core_presets import sys_io_interface as sii
from .core_presets import sys_io_exceptions as sie
from .core_presets import int_tabl_model as itm
from .drivers import LoaderConfigError


class FileIoInterface(abc.ABC):

    @abc.abstractmethod
    def load(self, read_params: typing.Any) -> typing.NoReturn:
        pass

    @abc.abstractmethod
    def save(self,
             data: typing.Any,
             write_params: typing.Any) -> typing.NoReturn:
        pass

    @abc.abstractmethod
    def close(self) -> typing.NoReturn:
        pass


class UnitOfWork(abc.ABC):

    @abc.abstractmethod
    def __enter__(self) -> typing.NoReturn:
        pass

    @abc.abstractmethod
    def __exit__(self,
                 exc_type: Exception,
                 exc_value: typing.Any,
                 traceback: typing.Any) -> typing.NoReturn:
        pass

    @property
    @abc.abstractmethod
    def port(self) -> typing.NoReturn:
        pass


def _starter(
        func: typing.Callable[..., typing.Generator]
        ) -> typing.Callable:

    @functools.wraps(func)
    def wrapper(*args, **kwargs) -> typing.Generator:

        generator = func(*args, **kwargs)
        generator.send(None)
        return generator

    return wrapper


class _ExcelPostLoader:

    def __init__(self) -> None:
        self._active = False
        self._max_idx = None

    @property
    def active(self) -> bool:
        return self._active

    def validate_first(
            self,
            line: typing.Tuple[typing.Any]
            ) -> None:
        if not self._first_item(line):
            msg = 'Value in first sheet cell not found'
            raise Exception(msg)

    def analyse_borders_equality(
            self,
            line: typing.Tuple[typing.Any],
            max_line_pos: int
            ) -> bool:
        self._max_idx = self._find_max_index(line)
        return self._max_idx == len(line) - 1

    def activate_postprocessing(self) -> None:
        self._active = True

    def deactivate_postprocessing(self) -> None:
        self._active = False

    def _first_item(
            self,
            line: typing.Tuple[typing.Any]
            ) -> bool:

        first = 0
        if line[first].value:
            return True
        return False

    def _last_item(
            self,
            line: typing.Tuple[typing.Any]
            ) -> bool:

        last = -1
        if line[last].value:
            return True
        return False

    def _find_max_index(
            self,
            line: typing.Tuple[typing.Any]
            ) -> int:

        x, dx = 0, len(line) - 1
        indexes = []

        while True:
            px = x + (dx - x) // 2
            if line[px].value:
                indexes.append(px)
            else:
                dx = px
                x = 0
                continue

            if x == px:
                return max(indexes)
            x = px

    def load(
            self,
            line: typing.Any
            ) -> typing.Tuple[typing.Any]:

        def _wrapped_loader(
                line: typing.Tuple[str]
                ) -> typing.Tuple[typing.Any]:

            checked_cells = []
            border = self._max_idx + 1

            for item in line[:border]:
                result = item.value is not None
                checked_cells.append(result)

            if any(checked_cells):
                return line[:border]
            else:
                raise StopIteration from None

        def _base_loader(
                line: typing.Tuple[str]
                ) -> typing.Tuple[typing.Any]:
            return line

        if self._active:
            return _wrapped_loader(line)
        return _base_loader(line)


class ExcelFileReader(sii.FileReaderInterface):

    def read(self, settings: typing.Any) -> oppxl.Workbook:
        book = oppxl.load_workbook(settings.path,
                                   read_only=settings.mode)
        sheet = self._set_active_sheet(book, settings.name)
        max_col, max_row = self._calculate_dims(sheet)

        need_postload_check = True
        postloader = None

        for row in sheet.iter_rows(
                max_row=max_row,
                max_col=max_col
                ):
            if need_postload_check:
                postloader = _ExcelPostLoader()
                postloader.validate_first(row)
                all_cells_valid = postloader.analyse_borders_equality(
                        row,
                        max_col
                        )
                if not all_cells_valid:
                    postloader.activate_postprocessing()
                need_postload_check = False

            if postloader is None or not postloader.active:
                yield row
            else:
                try:
                    yield postloader.load(row)
                except StopIteration:
                    break

        if postloader:
            postloader.deactivate_postprocessing()

    def _set_active_sheet(
            self,
            book: oppxl.Workbook,
            sheetname: str
            ) -> typing.Any:
        """
        Return active worksheet.
        """
        if sheetname in book.sheetnames:
            return book[sheetname]
        return book.active

    def _calculate_dims(
            self,
            sheet: typing.Any
            ) -> typing.Tuple[int, int]:

        return sheet.max_column, sheet.max_row


class TxtFileReader(sii.FileReaderInterface):
    """
    Test txt_reader implementation.
    """

    def read(
            self,
            settings: typing.Any
            ) -> typing.Generator:
        with open(settings.path) as file:
            for line in file:
                yield line


class ExcelFileWriter(sii.FileWriterInterface):

    def __init__(self) -> None:
        self._wb = None

    @_starter
    def write(
            self,
            settings: typing.Any,
            *,
            auto_closing: bool = False,
            ) -> None:

        sheet = None
        self._configure_workbook(settings)
        if self._wb:
            sheet = self._get_worksheet(settings.name)

        while True:
            try:
                line = yield
                if isinstance(line, list):
                    sheet.append(line)
            except StopIteration:
                self._save(settings.path)
        else:
            if auto_closing:
                self.close()

    def _configure_workbook(self, settings: typing.Any) -> None:
        if not isinstance(self._wb, oppxl.Workbook):
            try:
                self._wb = oppxl.load_workbook(
                                filename=settings.path,
                                )
            except Exception:
                pass
            finally:
                if self._wb is None:
                    self._wb = oppxl.Workbook(write_only=settings.mode)

    def _get_worksheet(self, name: str) -> typing.Any:
        if name not in self._wb.sheetnames:
            return self._wb.create_sheet(name)
        else:
            return self._wb[name]

    def _save(self, path: str) -> None:
        self._wb.save(path)

    def close(self) -> None:
        try:
            self._wb.close()
        except Exception:
            pass
        self._wb = None


class BaseFileIOAdapter(FileIoInterface):

    def __init__(
            self,
            loader: typing.Any,
            dumper: typing.Any,
            model: itm.TableSheetModel,
            ) -> None:
        self._loader = loader
        self._dumper = dumper
        self._model = model
        self._errors = collections.deque()

    @property
    def errors(self) -> typing.Any:
        """TODO -> while True & popleft from deque."""
        while self._errors:
            yield self._errors.popleft()

    def load(self, read_params: typing.Any) -> typing.Any:
        driver, loader = self._configure_load_sources(read_params)
        model = self._model.make_new_model()
        model.name = read_params.name

        while True:
            raw_data = loader.load()
            if raw_data is None:
                break

            try:
                if model.empty:
                    headers = driver.fetch_headers(raw_data)
                    model.add_headers(headers)
                else:
                    values = driver.fetch_values(raw_data)
                    if model.validate(values):
                        model.add_values(values)
            except sie.DriverError as e:
                self._errors.append(e)

        return model

    def _configure_load_sources(
                    self,
                    settings: typing.Any
                    ) -> typing.Optional[typing.Tuple[
                                sii.FileDriverInterface,
                                str,
                                ]]:
        try:
            self._loader.setup(settings)
            sources = self._loader.get_load_sources()
            self._validate_sources(sources)
            return sources
        except (LoaderConfigError, FileNotFoundError) as e:
            self._errors.append(e)
            self._loader.clean_setup()
            raise sie.AdapterError from e
        except Exception as e:
            self._loader.clean_setup()
            raise sie.AdapterError from e

    def _validate_sources(self, sources: typing.Tuple) -> None:
        if not all(sources) or len(sources) != 2:
            msg = f'Invalid sources: <{sources}>.'
            raise sie.AdapterError(msg)

    def save(self, model: itm.TableSheetModel,
             write_params: typing.Any) -> None:
        """Write by line."""
        sources = self._get_dump_sources(write_params)
        self._validate_sources(sources)
        driver, writer = sources

        for line in model.rows:
            try:
                compiled = driver.read(line)
                writer.send(compiled)
            except sie.DriverError as e:
                self._errors.append(e)

        writer.throw(StopIteration)
        writer.close()

    def _get_dump_sources(
                self,
                settings: typing.Any
                ) -> typing.Optional[typing.Tuple[
                            sii.FileDriverInterface,
                            str,
                            ]]:
        try:
            self._dumper.setup(settings)
            return self._dumper.get_dump_sources()
        except (LoaderConfigError, FileNotFoundError) as e:
            self._errors.append(e)
            self._dumper.clean_setup()
            raise sie.AdapterError from e
        except Exception as e:
            self._dumper.clean_setup()
            raise sie.AdapterError from e

    def close(self) -> None:
        self._loader.clean_setup()
        self._dumper.clean_setup()


class FileOperator(UnitOfWork):

    def __init__(self,
                 adapter: FileIoInterface,
                 logger: typing.Any) -> None:
        self._logger = logger
        self._adapter = adapter
        self._events = []

    @property
    def events(self) -> list:
        return []

    def __enter__(self) -> typing.Any:
        return self

    def __exit__(self,
                 exc_type: Exception,
                 exc_value: typing.Any,
                 traceback: typing.Any) -> None:
        if exc_type is not None:
            for err in self._adapter.errors:
                self._logger.critical(err)
            msg = f'{self.__class__.__name__} finished with '\
                  f'error: {exc_type}, msg: {exc_value}, '\
                  f'traceback: {traceback}.'
            self._logger.critical(msg)

    @property
    def port(self) -> FileIoInterface:
        return self._adapter
