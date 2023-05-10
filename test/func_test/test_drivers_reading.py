import pytest
import types
import collections
import logging

import openpyxl as oppxl

from services import drivers
from core.settings import settings


@pytest.fixture(scope='module')
def logger() -> logging.Logger:
    return logging.getLogger()


@pytest.fixture(scope='module')
def config() -> collections.OrderedDict:
    return settings.make_config()


@pytest.fixture(scope='module')
def multy_headers(config: collections.OrderedDict) -> None:
    m_headers = settings.MULTY_HEADERS_KEY
    p_map = settings.build_patterns_map(
            m_headers,
            config
            )
    return p_map


@pytest.fixture
def driver(
        logger: logger
        ) -> drivers.Driver:
    txt_compiler = drivers.TxtCompiler()
    driver = drivers.BaseDriver(
            logger,
            txt_compiler
            )
    return driver


@pytest.fixture
def excel_driver(logger: logging.Logger) -> drivers.ExcelDriver:
    excel_cmp = drivers.ExcelCompiler()
    return drivers.ExcelDriver(
            logger,
            excel_cmp
            )


@pytest.fixture(scope='module')
def rail_headers(config: collections.OrderedDict) -> None:
    r_headers = settings.RAIL_HEADERS_KEY
    r_map = settings.build_patterns_map(
            r_headers,
            config
            )
    return r_map


@pytest.fixture(scope='module')
def excel_wb() -> oppxl.Workbook:
    def_headers = ['POL', 'POD', 'vessel', 'smteu', 'bigteu', 'drop']
    wb = oppxl.Workbook()
    ws = wb.create_sheet('test')
    for idx, i in enumerate(def_headers, 1):
        cell = ws.cell(1, idx)
        cell.value = i.upper()
    return wb


@pytest.fixture(scope='module')
def valid_wb() -> oppxl.Workbook:
    def_headers = ['POL', 'POD', 'smteu', 'bigteu', 'drop', 'validity', 'info']
    wb = oppxl.Workbook()
    ws = wb.create_sheet('test')
    for idx, i in enumerate(def_headers, 1):
        cell = ws.cell(1, idx)
        cell.value = i.upper()
    return wb


def test_wb_fixt_creation(excel_wb: oppxl.Workbook) -> None:
    sheet = excel_wb.active
    for cell in sheet.iter_cols():
        assert cell.value is not None, f'Fail, val {cell.value}.'
        assert isinstance(cell.value, str), f'Invalid type: {type(cell.value)}'


def test_preset_fixture(
        multy_headers: types.MappingProxyType,
        ):
    assert list(multy_headers) != [], 'collection empty'


def test_txtcompiler_read_line(
        multy_headers: types.MappingProxyType,
        driver: drivers.Driver
        ) -> None:
    line = 'Shanghai-Vladivostok $2600/4800/4800 by HEUNG-A Excl DTHC $450/550'
    driver.headers_preset = multy_headers
    headers = driver.fetch_headers(line)
    assert headers == list(multy_headers), 'Error!!'


def test_txtcompiler_fetch_values(
        multy_headers: types.MappingProxyType,
        driver: drivers.Driver,
        ) -> None:
    line = 'Shanghai-Vladivostok $2600/4800/4800 by HEUNG-A Excl DTHC $450/550'
    driver.headers_preset = multy_headers
    values = driver.fetch_values(line)
    assert len(values) == len(multy_headers), 'unequal, incorrect reading.'


@pytest.mark.xfail(raises=drivers.DriverError)
def test_try_load_invalid_headers_from_wb(
        excel_wb: oppxl.Workbook,
        excel_driver: drivers.Driver,
        multy_headers: types.MappingProxyType
        ) -> None:
    sheet = excel_wb['test']
    excel_driver.headers_preset = multy_headers
    for row in sheet.iter_rows():
        excel_driver.fetch_headers(row)


def test_excel_driver_fetch_headers_and_values_correct(
        valid_wb: oppxl.Workbook,
        excel_driver: drivers.Driver,
        multy_headers: types.MappingProxyType
        ) -> None:
    sheet = valid_wb['test']
    excel_driver.headers_preset = multy_headers
    headers = None
    values = None
    for row in sheet.iter_rows():
        headers = excel_driver.fetch_headers(row)
        values = excel_driver.fetch_values(row)
    assert headers is not None, 'Headers is None'
    assert len(headers) - len(list(multy_headers)) == 1, 'bigger'
    assert values is not None, 'values None'
